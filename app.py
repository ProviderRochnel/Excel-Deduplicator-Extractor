import streamlit as st
import pandas as pd
import io
import os
import re
import numpy as np
import json
import requests
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

CRYPTO_AVAILABLE = True  # megapy gère le chiffrement en interne
try:
    from mega import Mega as _MegaTest
except ImportError:
    CRYPTO_AVAILABLE = False

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# CONFIGURATION ET CONSTANTES
# ============================================================================

# Configuration des traitements par catégorie
COLUMNS_CONFIG = {
    "Achat": {
        "extract_only": False,
        "description": "Optimisation des achats : extraction des données clés et suppression automatique des doublons de facturation.",
        "tooltip": "Idéal pour consolider vos bons de commande et éviter les doubles paiements.",
        "columns": ["Date", "Products", "Quantity ordered", "Price", "Order number"],
        "rename": {"Price": "Prix Achat HT", "Order number": "N° Facture"},
    },
    "Vente": {
        "extract_only": True,
        "description": "Préparation des données clients : extraction exhaustive pour analyse de facturation et logistique.",
        "tooltip": "Préparez vos fichiers pour la livraison et le suivi commercial.",
        "columns": [
            "Submit date", "Object code", "Object fullname", "Document number",
            "Delivery date", "Transport fees", "Product code", "Product",
            "Product type", "Type of article", "Quantity", "Sale unit code",
            "Position brutto value", "Position VAT rate", "PSA value", "PSA",
        ],
        "rename": {
            "Submit date": "Date/Heure Emission", "Object code": "Code Client", "Object fullname": "Nom Client",
            "Document number": "N° Facture", "Delivery date": "Date Livraison", "Transport fees": "Frais de port",
            "Product code": "Code Produit", "Product": "Désignation Produit", "Product type": "Type Produit",
            "Quantity": "Quantité Commandée", "Position brutto value": "Valeur Brutto", "Position VAT rate": "Taux de TVA",
            "PSA value": "Valeur PSA", "PSA": "PSA",
        },
        "extra_columns": ["N° Camion", "Chauffeur", "Vendeur", "Tournée"],
    },
    "Momo": {
        "extract_only": True,
        "description": "Traitement des flux mobiles (Momo) : conversion CSV/Excel et normalisation des transactions.",
        "tooltip": "Traitez vos rapports d'opérations mobiles en un clic.",
        "columns": ["Id", "Date", "Status", "Type", "From", "To","To name", "Amount", "Balance"],
        "rename": {"Id": "N° Identification", "From": "Provenance", "To name": "To handler name"},
        "extra_columns": ["Vendeur", "Compte", "Tournée"],
    },
    "Ristournes": {
        "extract_only": True,
        "description": "Calcul des ristournes : récapitulatif structuré des remises quotidiennes par magasin.",
        "tooltip": "Générez vos états de remises sur ventes en quelques secondes.",
        "columns": ["Code", "Store name", "Start date", "Document number", "Product code", "Product name", "Family", "CODE RABATES", "RISTOURNE"],
        "rename": {},
        "extra_columns": [],
    },
    "Mouvement de stock": {
        "extract_only": False,
        "description": "Mouvements de stock : analyse fine des opérations magasin (Chargement, Déchargement, etc.) et suppression automatique des doublons.",
        "tooltip": "Découpe automatiquement vos documents en Opérations et Comptes.",
        "columns": ["Product", "Change date", "Prev. quantity", "Cur. quantity", "Position record date", "Change source", "Document"],
        "rename": {
            "Product": "Product", "Change date": "Date Opération", "Prev. quantity": "Quantité Initiale",
            "Cur. quantity": "Quantité Résultante", "Position record date": "Journée", "Change source": "Référence", "Document": "Opération"
        },
        "extra_columns": ["Compte", "Responsable", "Camion", "Chauffeur"],
    },
    "STOCKS DES PRODUITS": {
        "extract_only": True,
        "description": "Inventaire financier : valorisation des stocks avec arrondis comptables et traduction de nature.",
        "tooltip": "Valorisez vos stocks au PA HT avec arrondis automatiques.",
        "columns": ["Code", "Name", "Category", "Price", "Qty", "Totalvalue"],
        "rename": {"Code": "Code", "Name": "Nom du Produit", "Category": "Nature", "Price": "Prix Achat HT", "Qty": "Qty", "Totalvalue": "Valeur au PA HT"},
        "extra_columns": ["Conditionnement", "PA TTC", "Valeur au PA TTC", "Prix Vente HT", "Prix de Vente 2%", "Valeur au PV TTC"],
    },
    "BASE CLIENTS": {
        "extract_only": True,
        "description": "Base CRM : extraction des paramètres clients (localisation, fiscalité, classification).",
        "tooltip": "Exportez votre base client avec ses coordonnées et typologies.",
        "columns": [
            "Code", "Name", "Street", "OBJ_CONTACT|BusinessPhone", "OBJ_PARAM_NC", 
            "OBJ_PARAM_ADDITIONAL_TAX", "OBJ_PARAM_DISTANCE", "OBJ_PARAM_TYPOLOGIE", 
            "OBJ_PARAM_ITINERAIRE", "Creation Date"
        ],
        "rename": {
            "Street": "Localisation", "OBJ_CONTACT|BusinessPhone": "Numéro Téléphone", "OBJ_PARAM_NC": "Numéro d'Identification Unique",
            "OBJ_PARAM_ADDITIONAL_TAX": "Statut Fiscal", "OBJ_PARAM_DISTANCE": "Position par rapport au Centre",
            "OBJ_PARAM_TYPOLOGIE": "TYPOLOGIE", "OBJ_PARAM_ITINERAIRE": "Classification",
        },
        "extra_columns": ["Route", "Tournée"],
    },
    "Inventaires": {
        "extract_only": True,
        "description": "Clôture journalière : nettoyage des stocks nuls et calcul des écarts d'inventaire.",
        "tooltip": "Supprime automatiquement les lignes sans stock pour un inventaire clair.",
        # FIX 1: "Valeur Stock" retiré des colonnes extraites (non présente dans le fichier source).
        # FIX 1b: extra_columns corrigé → colonne vide "Valeur Stock  PV TTC 2%" ajoutée en sortie.
        "columns": ["Date", "Code Article", "Conditionnement", "Nom Article", "Quantité Initiale", "Quantité Comptée"],
        "rename": {},
        "extra_columns": ["Valeur Stock  PV TTC 2%"],
    }
}

# Configuration des tâches quotidiennes avec icônes améliorées
DAILY_TASKS_CONFIG = {
    "Inventaire": {
        "category": "Inventaires",
        "icon": "📋",
        "description": "Clôture journalière des stocks",
        "priority": 1
    },
    "Données Magasin": {
        "category": "Mouvement de stock", 
        "icon": "🏪",
        "description": "Mouvements de stock du jour",
        "priority": 2
    },
    "Ristourne": {
        "category": "Ristournes",
        "icon": "💰", 
        "description": "Calcul des ristournes quotidiennes",
        "priority": 3
    },
    "Vente": {
        "category": "Vente",
        "icon": "👥",
        "description": "Facturation clients du jour", 
        "priority": 4
    },
    "Achat": {
        "category": "Achat",
        "icon": "🛒",
        "description": "Facturation achats du jour",
        "priority": 5
    }
}

# Constantes système
INDEX_SHEET_NAME = "Données"
TRACE_COLS = ["Fichier source", "Date d'import"]
BASE_MAGASIN_OPS = {"LR": "Chargement", "UC": "Déchargement", "DO": "Sorties directes de stock", "DI": "Entrées Directes de stock", "BO": "Achat ou Commande Appro", "DE": "Retour Emballages SABC"}
BASE_MAGASIN_COMPTES = {"S02004": "R1", "S02003": "R2", "S02005": "R3", "S02006": "R4"}
STOCKS_NATURES = {"BEER": "BIÈRES", "BG": "BOISSONS RAFRAICHISSANTE SANS ALCOOL", "AM": "ALCOOL MIX", "EAU": "EAUX MINERALES NATURELLES", "EMB": "PACKAGES"}

# Styles et couleurs
PRIMARY_COLOR = "#2E4057"
SUCCESS_COLOR = "#28a745"
WARNING_COLOR = "#ffc107"
DANGER_COLOR = "#dc3545"
LIGHT_COLOR = "#f8f9fa"


# ============================================================================
# UTILITAIRES SYSTÈME
# ============================================================================

class FileManager:
    """Gestion centralisée des opérations sur fichiers"""
    
    @staticmethod
    def get_index_folder() -> Path:
        """Crée et retourne le chemin du dossier DataHub_Index dans le dossier utilisateur"""
        user_folder = Path.home()
        index_folder = user_folder / "DataHub_Index"
        
        if index_folder.exists():
            if index_folder.is_dir():
                pass
            else:
                index_folder.unlink(missing_ok=True)
                index_folder.mkdir(exist_ok=True)
        else:
            index_folder.mkdir(exist_ok=True)
        
        return index_folder
    
    @staticmethod
    def save_index_locally(file_content: bytes, filename: str) -> bool:
        """Sauvegarde le contenu de l'index dans un fichier"""
        try:
            index_folder = FileManager.get_index_folder()
            file_path = index_folder / filename
            with open(file_path, "wb") as f:
                f.write(file_content)
            return True
        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde : {e}")
            return False
    
    @staticmethod
    def get_local_index(filename: str) -> Optional[bytes]:
        """Récupère le contenu d'un index depuis le dossier utilisateur"""
        try:
            index_folder = FileManager.get_index_folder()
            file_path = index_folder / filename
            if file_path.exists():
                with open(file_path, "rb") as f:
                    return f.read()
            return None
        except Exception as e:
            st.error(f"Erreur lors de la lecture : {e}")
            return None
    
    @staticmethod
    def list_local_indexes() -> List[str]:
        """Liste tous les index disponibles dans le dossier utilisateur"""
        try:
            index_folder = FileManager.get_index_folder()
            return [f.name for f in index_folder.glob("*.xlsx")]
        except Exception as e:
            st.error(f"Erreur lors de la liste des fichiers : {e}")
            return []
    
    @staticmethod
    def delete_index(filename: str) -> bool:
        """Supprime un index spécifique du dossier utilisateur"""
        try:
            index_folder = FileManager.get_index_folder()
            file_path = index_folder / filename
            if file_path.exists():
                file_path.unlink()
                return True
            return False
        except Exception as e:
            st.error(f"Erreur lors de la suppression : {e}")
            return False


class MegaUploader:
    """Upload vers Mega.nz via megapy (fork stable, compatible Streamlit Cloud).
    Installation : pip install megapy
    """

    def __init__(self, email: str, password: str):
        self.email = email
        self.password = password
        self._client = None

    def _get_client(self):
        if self._client is None:
            try:
                from mega import Mega
            except ImportError:
                raise RuntimeError(
                    "`megapy` n'est pas installé. "
                    "Ajoutez `megapy` dans requirements.txt et relancez l'application."
                )
            self._client = Mega().login(self.email, self.password)
        return self._client

    def login(self) -> bool:
        self._get_client()
        return True

    def upload(self, file_bytes: bytes, filename: str, parent_node=None) -> str:
        import tempfile, os
        client = self._get_client()
        with tempfile.NamedTemporaryFile(delete=False, suffix=f"_{filename}") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            file_node = client.upload(tmp_path, dest_filename=filename)
            return client.get_upload_link(file_node)
        finally:
            os.unlink(tmp_path)


class DailyTaskManager:
    """Gestion des tâches quotidiennes"""
    
    @staticmethod
    def get_task_file() -> Path:
        """Retourne le chemin du fichier de suivi des tâches"""
        return FileManager.get_index_folder() / "daily_tasks.json"
    
    @staticmethod
    def load_tasks() -> Dict[str, Any]:
        """Charge les tâches quotidiennes depuis le fichier JSON"""
        try:
            task_file = DailyTaskManager.get_task_file()
            if task_file.exists():
                with open(task_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {}
        except Exception:
            return {}
    
    @staticmethod
    def save_tasks(tasks_data: Dict[str, Any]) -> bool:
        """Sauvegarde les tâches quotidiennes dans le fichier JSON"""
        try:
            task_file = DailyTaskManager.get_task_file()
            with open(task_file, 'w', encoding='utf-8') as f:
                json.dump(tasks_data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            st.error(f"Erreur sauvegarde tâches : {e}")
            return False
    
    @staticmethod
    def get_today_key() -> str:
        """Retourne la clé pour aujourd'hui (format YYYY-MM-DD)"""
        return datetime.now().strftime("%Y-%m-%d")
    
    @staticmethod
    def is_task_completed(task_name: str) -> bool:
        """Vérifie si une tâche est complétée aujourd'hui"""
        tasks_data = DailyTaskManager.load_tasks()
        today_key = DailyTaskManager.get_today_key()
        return tasks_data.get(today_key, {}).get(task_name, False)
    
    @staticmethod
    def mark_task_completed(task_name: str) -> bool:
        """Marque une tâche comme complétée aujourd'hui"""
        tasks_data = DailyTaskManager.load_tasks()
        today_key = DailyTaskManager.get_today_key()
        
        if today_key not in tasks_data:
            tasks_data[today_key] = {}
        
        tasks_data[today_key][task_name] = True
        return DailyTaskManager.save_tasks(tasks_data)
    
    @staticmethod
    def get_progress() -> Tuple[int, int]:
        """Retourne le progrès quotidien (tâches complétées / total)"""
        completed = sum(1 for task_name in DAILY_TASKS_CONFIG.keys() 
                       if DailyTaskManager.is_task_completed(task_name))
        total = len(DAILY_TASKS_CONFIG)
        return completed, total


class ExcelStyler:
    """Gestion des styles Excel"""
    
    @staticmethod
    def apply_header_style(ws, nb_cols: int, row: int = 1) -> None:
        """Applique le style d'en-tête"""
        for col_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", start_color=PRIMARY_COLOR.replace("#", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def get_border() -> Border:
        """Retourne un bordure standard"""
        thin = Side(style="thin", color="BFBFBF")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    
    @staticmethod
    def apply_auto_width(ws) -> None:
        """Ajuste automatiquement la largeur des colonnes"""
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ============================================================================
# LOGIQUE DE TRAITEMENT DES DONNÉES
# ============================================================================

class DataProcessor:
    """Classe principale pour le traitement des données"""
    
    @staticmethod
    def filter_columns(df: pd.DataFrame, sheet_name: str, file_type: Optional[str], date_str: Optional[str] = None) -> Tuple[pd.DataFrame, List[str], List[str]]:
        """Filtre et transforme les colonnes selon le type de fichier"""
        if file_type is None or file_type not in COLUMNS_CONFIG:
            return df, [], list(df.columns)
        
        config = COLUMNS_CONFIG[file_type]
        expected = config["columns"]
        rename_map = config.get("rename", {})
        
        # Cas spécial pour STOCKS DES PRODUITS
        if file_type == "STOCKS DES PRODUITS" and "Code" not in df.columns and len(df.columns) >= 7:
            df["Code"] = df.iloc[:, 6]
        
        # Colonnes disponibles avec correspondance flexible pour INVENTAIRES
        if file_type == "Inventaires":
            # Correspondance flexible pour les noms de colonnes
            # Gère les variantes typographiques du fichier source (typos, espaces, retours à la ligne)
            column_mapping = {
                "Code Article": ["Code Article", "Code", "CodeArt", "Article Code", "Ref Article"],
                "Conditionnement": ["Conditionnement", "Conidtionnement", "Conidtionnement ","Cond", "Packaging", "Format"],
                "Nom Article": ["Nom Article", "Article", "Produit", "Désignation", "Nom Produit"],
                "Quantité Initiale": ["Quantité  Initiale", "Quantité Initiale", "Qte Initiale", "Stock Initial", "Qte Init"],
                "Quantité Comptée": ["Quantité Comptée", "Qte Comptée", "Stock Compté", "Qte Compte"],
            }
            
            # Injecter la date si elle n'est pas encore dans le DataFrame
            if "Date" not in df.columns and date_str:
                df.insert(0, "Date", date_str)

            available = []
            for expected_col in expected:
                found_col = None
                if expected_col in df.columns:
                    found_col = expected_col
                else:
                    for possible_name in column_mapping.get(expected_col, []):
                        if possible_name in df.columns:
                            found_col = possible_name
                            break
                
                if found_col:
                    if found_col != expected_col:
                        df = df.rename(columns={found_col: expected_col})
                    available.append(expected_col)
        else:
            available = [c for c in expected if c in df.columns]
        
        missing = [c for c in expected if c not in df.columns]
        df_out = df[available].copy()
        
        # Traitements spécifiques par type
        if file_type == "Inventaires":
            if not DataProcessor._validate_inventory_data(df_out):
                df_out = pd.DataFrame()
            else:
                df_out = DataProcessor._process_inventaires(df_out, date_str=date_str)
        elif file_type == "Mouvement de stock":
            df_out = DataProcessor._process_base_magasin(df_out)
        elif file_type == "STOCKS DES PRODUITS":
            df_out = DataProcessor._process_stocks(df_out)
        
        # Renommage et colonnes supplémentaires
        df_out = df_out.rename(columns=rename_map)
        for extra_col in config.get("extra_columns", []):
            if extra_col not in df_out.columns:
                df_out[extra_col] = ""
        
        return df_out, missing, list(df_out.columns)
    
    @staticmethod
    def _extract_inventaire_date(uploaded_file) -> Optional[str]:
        """Extrait la date/heure depuis la cellule A4 du fichier inventaire.
        Format attendu en A4: 'DATE & HEURE: 02/04/2026 06H33'
        """
        try:
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, data_only=True, read_only=True)
            ws = wb.active
            cell_a4 = ws["A4"].value
            wb.close()
            uploaded_file.seek(0)
            if cell_a4:
                cell_str = str(cell_a4).strip()
                match = re.search(r'(\d{2}/\d{2}/\d{4})\s+(\d{2}[Hh:]\d{2})', cell_str, re.IGNORECASE)
                if match:
                    date_part = match.group(1)
                    time_part = match.group(2).upper().replace(":", "H")
                    return f"{date_part} {time_part}"
                match_date = re.search(r'(\d{2}/\d{2}/\d{4})', cell_str)
                if match_date:
                    return match_date.group(1)
        except Exception:
            pass
        return None

    @staticmethod
    def _validate_inventory_data(df: pd.DataFrame) -> bool:
        """Valide que les données sont vraiment un inventaire"""
        if df.empty:
            return False
        
        required_cols = ["Code Article", "Nom Article"]
        inventory_cols = ["Quantité Initiale", "Quantité Comptée", "Valeur Ecart"]
        
        has_identification = any(col in df.columns for col in required_cols)
        if not has_identification:
            return False
        
        has_quantity = any(col in df.columns for col in inventory_cols)
        if not has_quantity:
            return False
        
        sample_data = df.head(10)
        if "Code Article" in df.columns:
            code_sample = sample_data["Code Article"].dropna().astype(str)
            monetary_pattern = code_sample.str.match(r'^\d{1,3}[\s,]?\d{3}[\s,]?\d{3}$')
            if monetary_pattern.sum() > len(code_sample) * 0.5:
                return False
        
        return True
    
    @staticmethod
    def _process_inventaires(df: pd.DataFrame, date_str: str) -> pd.DataFrame:
        """Traitement spécifique pour les inventaires.
        - Supprime les lignes où Quantité Initiale ET Quantité Comptée sont toutes deux à zéro.
        - Les lignes avec au moins une quantité non nulle sont CONSERVÉES (pas de perte de données).
        - Injecte la date extraite de A4 en colonne Date.
        """
        if "Quantité Initiale" in df.columns and "Quantité Comptée" in df.columns:
            qi = pd.to_numeric(df["Quantité Initiale"], errors="coerce").fillna(0)
            qc = pd.to_numeric(df["Quantité Comptée"], errors="coerce").fillna(0)
            # Conserver toute ligne où au moins une quantité est non nulle
            df = df[~((qi == 0) & (qc == 0))].copy()
        
        if "Date" not in df.columns:
            df.insert(0, "Date", date_str if date_str else "")
        else:
            df["Date"] = date_str if date_str else df["Date"]
            
        return df
    
    @staticmethod
    def _process_base_magasin(df: pd.DataFrame) -> pd.DataFrame:
        """Traitement spécifique pour la base magasin"""
        if "Position record date" in df.columns:
            df["Position record date"] = pd.to_datetime(df["Position record date"], errors='coerce').dt.strftime('%d/%m/%Y')
        
        if "Document" in df.columns:
            def extract_op_compte(doc_str):
                doc_str = str(doc_str)
                op_val = ""
                compte_val = ""
                for code, label in BASE_MAGASIN_OPS.items():
                    if code in doc_str:
                        op_val = label
                        break
                for code, label in BASE_MAGASIN_COMPTES.items():
                    if code in doc_str:
                        compte_val = label
                        break
                return pd.Series([op_val, compte_val])
            
            df[["Document", "Compte"]] = df["Document"].apply(extract_op_compte)
        else:
            df["Compte"] = ""
        
        dedup_cols = [c for c in ["Product", "Change date", "Prev. quantity", "Cur. quantity", "Change source", "Document"] if c in df.columns]
        if dedup_cols:
            df = df.drop_duplicates(subset=dedup_cols).reset_index(drop=True)
        
        return df
    
    @staticmethod
    def _process_stocks(df: pd.DataFrame) -> pd.DataFrame:
        """Traitement spécifique pour les stocks"""
        if "Category" in df.columns:
            df["Category"] = df["Category"].map(STOCKS_NATURES).fillna(df["Category"])
        
        for col in ["Price", "Totalvalue"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').apply(lambda x: np.ceil(x) if pd.notnull(x) else x)
        
        return df
    
    @staticmethod
    def load_data(uploaded_file, file_type: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """Charge les données depuis un fichier uploadé.
        Pour le module Inventaires, détecte automatiquement la ligne d'en-tête,
        ignore les lignes TOTAL (sous-totaux et total final) et s'arrête proprement.
        """
        filename = uploaded_file.name
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(uploaded_file, sep=None, engine='python')
            except:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, sep=',')
            return {"Données_CSV": df}
        else:
            # FIX 2 & 3: "FICHIER DES INVENTAIRES" renommé en "Inventaires" pour correspondre
            # à la clé COLUMNS_CONFIG sélectionnée dans l'interface.
            if file_type == "Inventaires":
                uploaded_file.seek(0)
                wb_inv = load_workbook(uploaded_file, data_only=True)
                result = {}
                KEYWORDS_INV = ["code article", "nom article", "quantité  initiale", "quantité initiale"]
                for sname in wb_inv.sheetnames:
                    ws_inv = wb_inv[sname]
                    header_row_idx = None
                    for i, row in enumerate(ws_inv.iter_rows(values_only=True), start=1):
                        vals = [str(v).replace("\n", " ").strip().lower() for v in row if v is not None]
                        if any(k in vals for k in KEYWORDS_INV):
                            header_row_idx = i
                            break
                    
                    if header_row_idx is not None:
                        headers_raw = [c.value for c in ws_inv[header_row_idx]]
                        valid_cols = [
                            (idx, str(v).replace("\n", " ").strip())
                            for idx, v in enumerate(headers_raw) if v is not None
                        ]
                        rows_data = []
                        for row in ws_inv.iter_rows(min_row=header_row_idx + 1, values_only=True):
                            # Ignorer les lignes TOTAL (sous-totaux par catégorie et total général)
                            row_str = " ".join([str(v) for v in row if v is not None]).upper()
                            if "TOTAL" in row_str:
                                continue
                            if not any(row):
                                continue
                            rows_data.append({h: row[idx] for idx, h in valid_cols})
                        df_clean = pd.DataFrame(rows_data)
                    else:
                        df_clean = pd.DataFrame(ws_inv.values)
                    result[sname] = df_clean
                wb_inv.close()
                uploaded_file.seek(0)
                return result
            else:
                return pd.read_excel(uploaded_file, sheet_name=None)
    
    @staticmethod
    def process_multiple_files(uploaded_files: List, file_type: Optional[str]) -> Tuple[bytes, List[Dict]]:
        """Traite plusieurs fichiers et génère le rapport"""
        all_results = []
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheets_written = 0
            for uploaded_file in uploaded_files:
                file_results = {"filename": uploaded_file.name, "sheets": []}
                try:
                    # FIX 2: extraction de la date calée sur le bon file_type "Inventaires"
                    inv_date_str = None
                    if file_type == "Inventaires" and not uploaded_file.name.endswith(".csv"):
                        inv_date_str = DataProcessor._extract_inventaire_date(uploaded_file)
                    
                    sheets_dict = DataProcessor.load_data(uploaded_file, file_type=file_type)
                    for sheet_name, df in sheets_dict.items():
                        total_rows = len(df)
                        df_proc, missing, final_cols = DataProcessor.filter_columns(df, sheet_name, file_type, date_str=inv_date_str)
                        
                        if file_type in COLUMNS_CONFIG:
                            extract_only = COLUMNS_CONFIG[file_type].get("extract_only", False)
                        else:
                            extract_only = False
                        
                        if extract_only:
                            unique_rows, duplicate_rows = len(df_proc), 0
                        else:
                            df_temp = df_proc.drop_duplicates()
                            unique_rows = len(df_temp)
                            duplicate_rows = len(df_proc) - unique_rows
                        
                        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', f"{uploaded_file.name[:20]}_{sheet_name[:10]}")
                        df_proc.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                        sheets_written += 1
                        
                        ws = writer.sheets[safe_sheet_name]
                        ExcelStyler.apply_header_style(ws, len(df_proc.columns))
                        ExcelStyler.apply_auto_width(ws)
                        
                        file_results["sheets"].append({
                            "name": sheet_name,
                            "total_rows": total_rows,
                            "unique_rows": unique_rows,
                            "duplicate_rows": duplicate_rows,
                            "missing_cols": missing
                        })
                except Exception as e:
                    file_results["error"] = str(e)
                
                all_results.append(file_results)
            
            if sheets_written == 0:
                pd.DataFrame({"Info": ["Aucune donnée traitée — vérifiez le format du fichier source."]}).to_excel(
                    writer, sheet_name="Résultat", index=False
                )
        
        return output.getvalue(), all_results


class IndexManager:
    """Gestion des index consolidés"""
    
    @staticmethod
    def merge_to_index(processed_bytes: bytes, category: str) -> Tuple[bool, str]:
        """Fusionne les nouvelles données avec l'index local existant"""
        try:
            filename = f"index_{category.replace(' ', '_')}.xlsx"
            new_data_io = io.BytesIO(processed_bytes)
            new_df_dict = pd.read_excel(new_data_io, sheet_name=None)
            
            new_df = pd.concat(new_df_dict.values(), ignore_index=True)
            
            existing_content = FileManager.get_local_index(filename)
            if existing_content:
                existing_df = pd.read_excel(io.BytesIO(existing_content))
                final_df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                final_df = new_df
            
            subset_cols = [c for c in final_df.columns if c not in TRACE_COLS]
            final_df = final_df.drop_duplicates(subset=subset_cols, keep='last')
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name=INDEX_SHEET_NAME)
                ws = writer.sheets[INDEX_SHEET_NAME]
                ExcelStyler.apply_header_style(ws, len(final_df.columns))
                ExcelStyler.apply_auto_width(ws)
            
            if FileManager.save_index_locally(output.getvalue(), filename):
                return True, filename
            return False, ""
        except Exception as e:
            st.error(f"Erreur fusion index : {e}")
            return False, ""


# ============================================================================
# COMPOSANTS UI
# ============================================================================

class UIComponents:
    """Gestion des composants d'interface Streamlit"""
    
    @staticmethod
    def setup_page_config() -> None:
        """Configure les paramètres de la page"""
        st.set_page_config(
            page_title="DataHub Pro | Excel Optimizer",
            page_icon="📊",
            layout="wide",
            initial_sidebar_state="expanded"
        )
    
    @staticmethod
    def apply_custom_styles() -> None:
        """Applique le CSS personnalisé"""
        st.markdown(f"""
        <style>
            .main {{ background-color: #f0f2f6; }}
            .stApp {{ background-color: #f0f2f6; }}
            .stButton>button {{ 
                border-radius: 12px; 
                height: 3.5em; 
                font-weight: 700; 
                transition: all 0.3s; 
            }}
            .stButton>button:hover {{ 
                transform: translateY(-2px); 
                box-shadow: 0 4px 12px rgba(0,0,0,0.1); 
            }}
            .step-card {{ 
                background-color: #ffffff; 
                padding: 25px; 
                border-radius: 15px; 
                margin-bottom: 25px; 
                box-shadow: 0 4px 6px rgba(0,0,0,0.05); 
                border-top: 5px solid {PRIMARY_COLOR}; 
            }}
            .step-header {{ 
                display: flex; 
                align-items: center; 
                margin-bottom: 15px; 
            }}
            .step-number {{ 
                background-color: {PRIMARY_COLOR}; 
                color: white; 
                border-radius: 50%; 
                width: 30px; 
                height: 30px; 
                display: flex; 
                align-items: center; 
                justify-content: center; 
                font-weight: bold; 
                margin-right: 15px; 
            }}
            .step-title {{ 
                color: {PRIMARY_COLOR}; 
                font-size: 1.3em; 
                font-weight: bold; 
            }}
            .stat-card {{
                background: white;
                padding: 20px;
                border-radius: 12px;
                text-align: center;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                border: 1px solid #eee;
            }}
            .stat-val {{
                font-size: 1.8em;
                font-weight: 800;
                color: {PRIMARY_COLOR};
            }}
            .stat-label {{
                color: #6c757d;
                font-size: 0.9em;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            .task-card {{
                background: white;
                padding: 15px 20px;
                border-radius: 12px;
                margin-bottom: 10px;
                display: flex;
                align-items: center;
                border-left: 5px solid {PRIMARY_COLOR};
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            }}
            .task-card-completed {{
                border-left-color: {SUCCESS_COLOR};
                background-color: #f8fff9;
            }}
            .task-icon {{
                font-size: 1.5em;
                margin-right: 20px;
            }}
            .task-content {{
                flex-grow: 1;
            }}
            .task-title {{
                font-weight: 700;
                color: {PRIMARY_COLOR};
            }}
            .task-desc {{
                color: #6c757d;
                font-size: 0.9em;
            }}
            .task-badge {{
                padding: 4px 10px;
                border-radius: 20px;
                font-size: 0.75em;
                font-weight: bold;
                text-transform: uppercase;
                margin-top: 8px;
                display: inline-block;
            }}
            .badge-todo {{ background-color: #ffeeba; color: #856404; }}
            .badge-done {{ background-color: #d4edda; color: #155724; }}
            
            .footer {{ 
                text-align: center; 
                color: #adb5bd; 
                padding: 40px 0; 
                font-size: 0.85em; 
            }}
        </style>
        """, unsafe_allow_html=True)
    
    @staticmethod
    def render_sidebar() -> None:
        """Affiche la barre latérale"""
        with st.sidebar:
            st.markdown(f"<h2 style='color:{PRIMARY_COLOR};'>Centre de Contrôle</h2>", unsafe_allow_html=True)
            
            if st.button("📊 Traitement des fichiers", use_container_width=True):
                st.session_state.page = "Traitement"
                st.rerun()
            if st.button("📅 Tâches Quotidiennes", use_container_width=True):
                st.session_state.page = "Taches"
                st.rerun()
            if st.button("📚 Bibliothèque des Index", use_container_width=True):
                st.session_state.page = "Index"
                st.rerun()
            
            st.markdown("---")
            if st.button("🔄 Réinitialiser la session", use_container_width=True):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()


class DataHubApp:
    """Application principale DataHub"""
    
    def __init__(self):
        self.processor = DataProcessor()
        self.index_manager = IndexManager()
        self.task_manager = DailyTaskManager()
        self.file_manager = FileManager()
        self.ui = UIComponents()
    
    def initialize_session(self) -> None:
        """Initialise l'état de la session"""
        if 'results' not in st.session_state:
            st.session_state.results = None
        if 'processed_data' not in st.session_state:
            st.session_state.processed_data = None
        if 'current_file_type' not in st.session_state:
            st.session_state.current_file_type = None
        if 'page' not in st.session_state:
            st.session_state.page = "Traitement"
        
        try:
            FileManager.get_index_folder()
        except Exception as e:
            st.error(f"Erreur lors de la création du dossier de stockage : {e}")
    
    def render_processing_page(self) -> None:
        """Affiche la page de traitement"""
        st.markdown("<h1>Traitement Intelligent</h1>", unsafe_allow_html=True)
        
        col_main1, col_main2 = st.columns([1, 1])
        
        with col_main1:
            st.markdown(f"""
            <div class="step-card">
                <div class="step-header">
                    <div class="step-number">1</div>
                    <div class="step-title">Configuration</div>
                </div>
            """, unsafe_allow_html=True)
            
            file_type = st.selectbox(
                "Module métier", 
                options=list(COLUMNS_CONFIG.keys()),
                index=None,
                placeholder="Sélectionnez une catégorie..."
            )
            
            if file_type:
                st.markdown(f"*{COLUMNS_CONFIG[file_type]['description']}*")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col_main2:
            st.markdown(f"""
            <div class="step-card">
                <div class="step-header">
                    <div class="step-number">2</div>
                    <div class="step-title">Source de Données</div>
                </div>
            """, unsafe_allow_html=True)
            
            uploaded_files = st.file_uploader(
                "Glissez vos fichiers Excel ou CSV ici", 
                type=["xlsx", "csv"], 
                accept_multiple_files=True
            )
            
            if uploaded_files:
                button_disabled = not file_type
                if button_disabled:
                    st.warning("Veuillez sélectionner une catégorie de traitement avant de lancer l'analyse.")
                
                if st.button("Lancer l'analyse", use_container_width=True, type="primary", disabled=button_disabled):
                    with st.spinner("Traitement en cours..."):
                        try:
                            p_data, results = self.processor.process_multiple_files(uploaded_files, file_type)
                            st.session_state.processed_data = p_data
                            st.session_state.results = results
                            st.session_state.current_file_type = file_type
                            
                            if file_type:
                                for task_name, task_config in DAILY_TASKS_CONFIG.items():
                                    if task_config["category"] == file_type:
                                        if self.task_manager.mark_task_completed(task_name):
                                            st.success(f"Tâche quotidienne '{task_name}' marquée comme complétée !")
                                        break
                            
                            st.balloons()
                        except Exception as e:
                            st.error(f"Erreur : {str(e)}")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        if st.session_state.results:
            self._render_results_section()
    
    def _render_results_section(self) -> None:
        """Affiche la section des résultats"""
        st.markdown(f"""
        <div class="step-card">
            <div class="step-header">
                <div class="step-number">3</div>
                <div class="step-title">Tableau de Bord & Actions</div>
            </div>
        """, unsafe_allow_html=True)
        
        total_in = sum(sum(s['total_rows'] for s in f['sheets']) for f in st.session_state.results)
        total_out = sum(sum(s['unique_rows'] for s in f['sheets']) for f in st.session_state.results)
        
        s1, s2, s3 = st.columns(3)
        with s1:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-val">{total_in}</div>
                <div class="stat-label">Lignes Entrantes</div>
            </div>
            """, unsafe_allow_html=True)
        
        with s2:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-val" style="color:{SUCCESS_COLOR};">{total_out}</div>
                <div class="stat-label">Lignes Uniques</div>
            </div>
            """, unsafe_allow_html=True)
        
        with s3:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-val" style="color:{DANGER_COLOR};">{total_in - total_out}</div>
                <div class="stat-label">Doublons Éliminés</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        c_left, c_right = st.columns([2, 1])
        
        with c_left:
            st.markdown("### Aperçu")
            tab_data = []
            for f in st.session_state.results:
                if "error" in f:
                    continue
                for s in f['sheets']:
                    tab_data.append({
                        "Fichier": f["filename"],
                        "Feuille": s.get("name", ""),
                        "Lignes unique": s['unique_rows'],
                    })
            if tab_data:
                st.dataframe(pd.DataFrame(tab_data), use_container_width=True, hide_index=True)
            else:
                st.warning("Aucune donnée à afficher.")
        
        with c_right:
            st.markdown("### Actions")
            suffix = "Standard" if st.session_state.current_file_type is None else st.session_state.current_file_type.replace(" ", "_")
            st.download_button(
                "Exporter Fichier Traité", 
                data=st.session_state.processed_data, 
                file_name=f"Export_{suffix}.xlsx", 
                use_container_width=True
            )
            
            if st.session_state.current_file_type:
                if st.button(f"Fusionner vers Index {st.session_state.current_file_type}", use_container_width=True):
                    success, fname = self.index_manager.merge_to_index(st.session_state.processed_data, st.session_state.current_file_type)
                    if success:
                        st.success(f"Données fusionnées dans `{fname}`")
                        st.session_state.fusion_done = True
                    else:
                        st.error("Erreur lors de la fusion de l'index.")

                if st.session_state.get("fusion_done"):
                    if st.button("Voir la bibliothèque →", use_container_width=True):
                        st.session_state.page = "Index"
                        st.session_state.fusion_done = False
                        st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)

    def render_daily_tasks_page(self) -> None:
        """Affiche la page des tâches quotidiennes"""
        st.markdown("<h1>Suivi des Tâches Quotidiennes</h1>", unsafe_allow_html=True)
        
        completed, total = self.task_manager.get_progress()
        progress_percent = (completed / total) * 100 if total > 0 else 0
        today_str = datetime.now().strftime("%A %d %B %Y")
        
        col_head1, col_head2 = st.columns([2, 1])
        with col_head1:
            st.markdown(f"### {today_str}")
            st.markdown(f"Complétion : **{completed} sur {total} tâches**")
        with col_head2:
            st.markdown(f"<h2 style='text-align:right; color:{SUCCESS_COLOR if completed==total else PRIMARY_COLOR};'>{progress_percent:.0f}%</h2>", unsafe_allow_html=True)
        
        st.progress(progress_percent / 100)
        st.markdown("<br>", unsafe_allow_html=True)
        
        for task_name, task_config in sorted(DAILY_TASKS_CONFIG.items(), key=lambda x: x[1]["priority"]):
            is_completed = self.task_manager.is_task_completed(task_name)
            
            status_class = "task-card-completed" if is_completed else ""
            badge_class = "badge-done" if is_completed else "badge-todo"
            badge_text = "Complété" if is_completed else "À faire"
            
            st.markdown(f"""
            <div class="task-card {status_class}">
                <div class="task-icon">{task_config['icon']}</div>
                <div class="task-content">
                    <div class="task-title">{task_name}</div>
                    <div class="task-desc">{task_config['description']}</div>
                    <div class="task-badge {badge_class}">{badge_text}</div>
                </div>
                <div style="font-size: 1.5em;">
                    {'✅' if is_completed else '⏳'}
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        col_info1, col_info2 = st.columns([1, 1])
        with col_info1:
            if completed == total:
                st.success("Félicitations ! Toutes les tâches sont terminées.")
            else:
                st.info(f"Il vous reste {total - completed} tâches pour clôturer la journée.")
        
        with col_info2:
            if st.button("➡️ Aller au Traitement", use_container_width=True, type="primary"):
                st.session_state.page = "Traitement"
                st.rerun()

    def _render_mega_config(self) -> None:
        """Affiche le panneau de configuration Mega dans la sidebar"""
        with st.sidebar:
            st.markdown("---")
            st.markdown(f"<h3 style='color:{PRIMARY_COLOR};'>☁️ Mega.nz</h3>", unsafe_allow_html=True)
            
            if not CRYPTO_AVAILABLE:
                st.warning("⚠️ `pycryptodome` manquant. Ajoutez-le dans requirements.txt pour activer l'envoi Mega.")
                return
            
            with st.expander("🔐 Identifiants Mega", expanded=False):
                mega_email = st.text_input("Email Mega", key="mega_email", placeholder="user@exemple.com")
                mega_pwd   = st.text_input("Mot de passe", key="mega_pwd", type="password")
                if mega_email and mega_pwd:
                    st.session_state["mega_creds"] = (mega_email, mega_pwd)
                    st.success("Identifiants enregistrés pour cette session.")
                elif "mega_creds" not in st.session_state:
                    st.info("Renseignez vos identifiants pour activer l'envoi vers Mega.")

    def _upload_to_mega(self, file_bytes: bytes, filename: str) -> Optional[str]:
        """Lance l'upload d'un fichier vers Mega et retourne le lien public."""
        creds = st.session_state.get("mega_creds")
        if not creds:
            st.error("Identifiants Mega non configurés. Renseignez-les dans la barre latérale.")
            return None
        
        email, pwd = creds
        try:
            uploader = MegaUploader(email, pwd)
            with st.spinner(f"Connexion à Mega…"):
                uploader.login()
            with st.spinner(f"Envoi de {filename} vers Mega…"):
                link = uploader.upload(file_bytes, filename)
            return link
        except RuntimeError as e:
            st.error(f"Erreur Mega : {e}")
            return None
        except Exception as e:
            st.error(f"Erreur inattendue lors de l'upload : {e}")
            return None

    def render_index_page(self) -> None:
        """Affiche la page de la bibliothèque d'index"""
        st.markdown("<h1>📚 Bibliothèque des Index</h1>", unsafe_allow_html=True)
        self._render_mega_config()
        
        try:
            indexes = self.file_manager.list_local_indexes()
        except Exception as e:
            st.error(f"Erreur d'accès : {e}")
            indexes = []
        
        if not indexes:
            st.warning("Aucun index trouvé. Traitez des fichiers pour commencer à construire votre base de données.")
        else:
            for idx_name in sorted(indexes):
                with st.expander(f"📄 {idx_name}", expanded=False):
                    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
                    with c1:
                        cat = idx_name.replace('index_', '').replace('.xlsx', '').replace('_', ' ')
                        st.markdown(f"**Catégorie** : {cat}")
                    
                    content = self.file_manager.get_local_index(idx_name)

                    with c2:
                        if content:
                            st.download_button(
                                "⬇️ Télécharger", content, idx_name,
                                key=f"dl_{idx_name}", use_container_width=True
                            )
                    
                    with c3:
                        mega_enabled = CRYPTO_AVAILABLE and st.session_state.get("mega_creds") is not None
                        if st.button(
                            "☁️ Envoyer Mega",
                            key=f"mega_{idx_name}",
                            use_container_width=True,
                            disabled=not (content and mega_enabled),
                            help="Configurez vos identifiants Mega dans la barre latérale" if not mega_enabled else None
                        ):
                            link = self._upload_to_mega(content, idx_name)
                            if link:
                                st.success(f"✅ Fichier envoyé !")
                                st.markdown(f"[🔗 Ouvrir sur Mega]({link})", unsafe_allow_html=False)
                                # Mémoriser le lien pour cet index
                                st.session_state[f"mega_link_{idx_name}"] = link
                        
                        # Afficher le dernier lien si disponible
                        if f"mega_link_{idx_name}" in st.session_state:
                            st.markdown(
                                f"[🔗 Dernier lien]({st.session_state[f'mega_link_{idx_name}']})",
                                unsafe_allow_html=False
                            )

                    with c4:
                        if st.button("🗑️ Supprimer", key=f"del_{idx_name}", use_container_width=True):
                            if self.file_manager.delete_index(idx_name):
                                st.rerun()

    def run(self) -> None:
        """Point d'entrée principal de l'application"""
        self.ui.setup_page_config()
        self.ui.apply_custom_styles()
        self.initialize_session()
        self.ui.render_sidebar()
        
        if st.session_state.page == "Traitement":
            self.render_processing_page()
        elif st.session_state.page == "Taches":
            self.render_daily_tasks_page()
        elif st.session_state.page == "Index":
            self.render_index_page()
        
        st.markdown('<div class="footer">DataHub Pro v3.1 | Stockage Local | © 2024</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    app = DataHubApp()
    app.run()
